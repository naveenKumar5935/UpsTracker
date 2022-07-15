import openpyxl
from openpyxl import Workbook
import requests
import pandas as pd
from bs4 import BeautifulSoup

def searchAndSave(r):

    url = 'https://www.bing.com/packagetrackingv2?packNum='+trackingNo+'&carrier=UPS&FORM=PCKTR1'
    url1 = 'https://httpbin.org/headers'

    response = requests.get(url=url)

    parse = BeautifulSoup(response.content,'html.parser')
    status = parse.find('div', {"class":"b_focusTextSmall"})
    if status is not None:
        print(status.text)
        print(sheet.max_row)
        row = "B"+str(r)
        if status.text.split(":")[0] != "Delivered":
            sheet[row].value = "UnDelivered"
        else:
            sheet[row].value = status.text.split(":")[0]

#Variables-------
trackingNoRow = "A"
trackingNoRowStart = 2

statusRow = "B"
workbookName = "trackingUps.xlsx"
sheetName = 'Sheet1'



workbook = openpyxl.load_workbook(workbookName)
sheet = workbook[sheetName]

for row in range(trackingNoRowStart,sheet.max_row+1):

    put = trackingNoRow+str(row)
    print(put)
    if sheet[put].value is not None:
        trackingNo = sheet[put].value
        searchAndSave(row)


workbook.save("trackingUps.xlsx")
#print(response.content)


